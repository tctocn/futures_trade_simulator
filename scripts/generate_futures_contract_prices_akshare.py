#!/usr/bin/env python3
"""基于 AKShare 生成期货合约价导入文件（交互式输入，多合约支持）。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import akshare as ak
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
TEMPLATE_PATH = PROJECT_ROOT / "template" / "期货合约价导入.xlsx"

HEADERS = ["合约代码", "结算价", "收盘价", "昨日结算价", "结算日期"]


@dataclass(frozen=True)
class UserInput:
    contracts: list[str]
    start_date: date
    end_date: date
    output: str


def prompt_with_default(prompt_text: str, default: str) -> str:
    text = input(f"{prompt_text} [{default}]: ").strip()
    return text if text else default


def parse_contracts(raw: str) -> list[str]:
    items = [x.strip().upper() for x in raw.split(",") if x.strip()]
    if not items:
        raise ValueError("合约代码不能为空")

    pat = re.compile(r"^[A-Z]{1,2}\d{3,4}$")
    result: list[str] = []
    for item in items:
        if not pat.match(item):
            raise ValueError(f"合约格式错误: {item}，示例: RB2610")
        if item not in result:
            result.append(item)
    return result


def read_user_input() -> UserInput:
    today = date.today()
    default_end = today
    default_start = today - timedelta(days=365)

    print("请输入参数（直接回车使用默认值）")
    contracts_text = prompt_with_default("合约代码(英文逗号分隔)", "RB2610,AG2612,CU2609")
    start_text = prompt_with_default("开始日期(YYYY-MM-DD)", default_start.strftime("%Y-%m-%d"))
    end_text = prompt_with_default("结束日期(YYYY-MM-DD)", default_end.strftime("%Y-%m-%d"))
    output = input("输出xlsx路径(留空则自动dist+时间戳): ").strip()

    start = datetime.strptime(start_text, "%Y-%m-%d").date()
    end = datetime.strptime(end_text, "%Y-%m-%d").date()
    if start > end:
        raise ValueError("开始日期不能晚于结束日期")

    return UserInput(
        contracts=parse_contracts(contracts_text),
        start_date=start,
        end_date=end,
        output=output,
    )


def build_output_path(raw_output: str) -> Path:
    if raw_output:
        out = Path(raw_output)
        if out.suffix.lower() != ".xlsx":
            out = out.with_suffix(".xlsx")
        return out if out.is_absolute() else (Path.cwd() / out).resolve()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return PROJECT_ROOT / "dist" / f"期货合约价导入_{ts}.xlsx"


def to_float_str(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        return f"{float(value):.2f}"
    except Exception:
        return ""


def fetch_contract_rows(contract: str, start_date: date, end_date: date) -> list[dict[str, str]]:
    df = ak.futures_zh_daily_sina(symbol=contract)
    if df.empty:
        return []

    work = df.copy()
    work["date"] = pd.to_datetime(work["date"]).dt.date
    work = work.sort_values("date").reset_index(drop=True)
    work["pre_settle"] = work["settle"].shift(1)

    mask = (work["date"] >= start_date) & (work["date"] <= end_date)
    work = work.loc[mask]

    rows: list[dict[str, str]] = []
    for _, r in work.iterrows():
        rows.append(
            {
                "合约代码": contract,
                "结算价": to_float_str(r.get("settle")),
                "收盘价": to_float_str(r.get("close")),
                "昨日结算价": to_float_str(r.get("pre_settle")),
                "结算日期": r.get("date").strftime("%Y-%m-%d") if pd.notna(r.get("date")) else "",
            }
        )
    return rows


def write_to_template(rows: list[dict[str, str]], output_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"模板不存在: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]

    actual_headers = [ws.cell(1, i).value or "" for i in range(1, len(HEADERS) + 1)]
    if actual_headers != HEADERS:
        raise ValueError(f"模板表头不匹配: {actual_headers}")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, row in enumerate(rows, start=2):
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(idx, col, row.get(header, ""))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    user = read_user_input()
    print("正在通过 AKShare 拉取合约价格，请稍候...")

    all_rows: list[dict[str, str]] = []
    empty_contracts: list[str] = []

    for contract in user.contracts:
        rows = fetch_contract_rows(contract, user.start_date, user.end_date)
        if not rows:
            empty_contracts.append(contract)
            continue
        all_rows.extend(rows)

    if not all_rows:
        raise ValueError("指定合约在时间范围内无可用数据")

    all_rows.sort(key=lambda x: (x["结算日期"], x["合约代码"]))
    output_path = build_output_path(user.output)
    write_to_template(all_rows, output_path)

    print(f"已生成 {len(all_rows)} 条合约价格记录: {output_path}")
    if empty_contracts:
        print(f"以下合约在时间范围内无数据，已跳过: {', '.join(empty_contracts)}")


if __name__ == "__main__":
    main()
