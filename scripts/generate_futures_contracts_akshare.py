#!/usr/bin/env python3
"""基于 AKShare 生成期货合约导入文件（交互式输入，多合约支持）。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import akshare as ak
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
TEMPLATE_PATH = PROJECT_ROOT / "template" / "期货合约导入.xlsx"

HEADERS = [
    "合约代码",
    "合约名称",
    "产品类型",
    "合约月份",
    "衍生品代码",
    "交易所",
    "币种",
    "合约乘数",
    "合约上市日",
    "最后交易日",
    "最小变动价格",
    "临近交割月结束日",
    "保证金比例%",
    "计量单位",
    "标的资产",
]

REMOVED_HEADERS = {
    "过期",
    "类型",
    "标的资产性质",
    "组合类型",
    "行权价",
    "涨跌",
}

PRODUCT_EXCHANGE_FALLBACK = {
    "AG": "上海期货交易所",
    "AL": "上海期货交易所",
    "AU": "上海期货交易所",
    "BU": "上海期货交易所",
    "CU": "上海期货交易所",
    "FU": "上海期货交易所",
    "HC": "上海期货交易所",
    "NI": "上海期货交易所",
    "PB": "上海期货交易所",
    "RB": "上海期货交易所",
    "RU": "上海期货交易所",
    "SN": "上海期货交易所",
    "SP": "上海期货交易所",
    "SS": "上海期货交易所",
    "WR": "上海期货交易所",
    "A": "大连商品交易所",
    "B": "大连商品交易所",
    "C": "大连商品交易所",
    "CS": "大连商品交易所",
    "EB": "大连商品交易所",
    "EG": "大连商品交易所",
    "I": "大连商品交易所",
    "J": "大连商品交易所",
    "JD": "大连商品交易所",
    "JM": "大连商品交易所",
    "L": "大连商品交易所",
    "LH": "大连商品交易所",
    "M": "大连商品交易所",
    "P": "大连商品交易所",
    "PG": "大连商品交易所",
    "PP": "大连商品交易所",
    "V": "大连商品交易所",
    "Y": "大连商品交易所",
}


@dataclass(frozen=True)
class UserInput:
    contracts: list[str]
    query_date: str
    output: str


def prompt_with_default(prompt_text: str, default: str) -> str:
    text = input(f"{prompt_text} [{default}]: ").strip()
    return text if text else default


def parse_contracts(raw: str) -> list[str]:
    items = [x.strip().upper() for x in raw.split(",") if x.strip()]
    if not items:
        raise ValueError("合约代码不能为空")
    pat = re.compile(r"^([A-Z]{1,2})(\d{4})$")
    result: list[str] = []
    for item in items:
        if not pat.match(item):
            raise ValueError(f"合约格式错误: {item}，示例: RB2610")
        if item not in result:
            result.append(item)
    return result


def read_user_input() -> UserInput:
    print("请输入参数（直接回车使用默认值）")
    contracts_text = prompt_with_default("合约代码(英文逗号分隔)", "RB2610,AG2612,CU2609")
    query_date = prompt_with_default("查询交易日(YYYYMMDD)", date.today().strftime("%Y%m%d"))
    output = input("输出xlsx路径(留空则自动dist+时间戳): ").strip()
    datetime.strptime(query_date, "%Y%m%d")
    return UserInput(contracts=parse_contracts(contracts_text), query_date=query_date, output=output)


def normalize_date_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def first_number(text: str) -> str:
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    return m.group(0) if m else ""


def parse_multiplier(text: str) -> str:
    # 例如: 10吨/手 -> 10
    return first_number(text)


def parse_unit(text: str) -> str:
    # 例如: 10吨/手 -> 吨
    m = re.search(r"\d+(?:\.\d+)?\s*([^/]+?)\s*/\s*手", text)
    if m:
        return m.group(1).strip()
    m2 = re.search(r"/([\u4e00-\u9fa5A-Za-z]+)$", text)
    return m2.group(1) if m2 else ""


def parse_margin_percent(text: str) -> str:
    # 例如: 合约价值的13% -> 13
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    return m.group(1) if m else ""


def build_output_path(raw_output: str) -> Path:
    if raw_output:
        out = Path(raw_output)
        if out.suffix.lower() != ".xlsx":
            out = out.with_suffix(".xlsx")
        return out if out.is_absolute() else (Path.cwd() / out).resolve()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return PROJECT_ROOT / "dist" / f"期货合约导入_{ts}.xlsx"


def fetch_exchange_base_info(query_date: str) -> dict[str, dict[str, Any]]:
    mapping: dict[str, dict[str, Any]] = {}
    apis = [
        (ak.futures_contract_info_shfe, {"date": query_date}),
        (ak.futures_contract_info_ine, {"date": query_date}),
        (ak.futures_contract_info_cffex, {"date": query_date}),
        (ak.futures_contract_info_gfex, {}),
    ]
    for fn, kwargs in apis:
        try:
            df = fn(**kwargs)
        except Exception:
            continue
        if "合约代码" not in df.columns:
            continue
        for _, row in df.iterrows():
            code = str(row.get("合约代码", "")).strip().lower()
            if code:
                mapping[code] = row.to_dict()
    return mapping


def fetch_comm_info_map() -> dict[str, dict[str, Any]]:
    df = ak.futures_comm_info(symbol="所有")
    mapping: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        code = str(row.get("合约代码", "")).strip().lower()
        if code:
            mapping[code] = row.to_dict()
    return mapping


def fetch_contract_detail_map(contract: str) -> dict[str, str]:
    try:
        df = ak.futures_contract_detail(symbol=contract)
    except Exception:
        return {}
    detail: dict[str, str] = {}
    for _, row in df.iterrows():
        k = str(row.get("item", "")).strip()
        v = str(row.get("value", "")).strip()
        if k:
            detail[k] = v
    return detail


def contract_month_to_date(month_yy_mm: str) -> str:
    # 例如: 2607 -> 2026-07-01
    if not re.fullmatch(r"\d{4}", month_yy_mm):
        return month_yy_mm
    year = 2000 + int(month_yy_mm[:2])
    month = int(month_yy_mm[2:])
    if not (1 <= month <= 12):
        return month_yy_mm
    return f"{year:04d}-{month:02d}-01"


def build_row(
    contract: str,
    base_info: dict[str, dict[str, Any]],
    comm_map: dict[str, dict[str, Any]],
    detail_map: dict[str, dict[str, str]],
) -> dict[str, str]:
    code_lower = contract.lower()
    prod_code = re.match(r"^([A-Z]{1,2})", contract).group(1)  # guaranteed by parse
    month = re.search(r"(\d{4})$", contract).group(1)

    comm = comm_map.get(code_lower, {})
    detail = detail_map.get(contract, {})
    base = base_info.get(code_lower, {})

    exchange_name = (
        str(comm.get("交易所名称", ""))
        or str(detail.get("上市交易所", ""))
        or PRODUCT_EXCHANGE_FALLBACK.get(prod_code, "")
    )
    underlying = str(detail.get("交易品种", ""))
    contract_name = str(comm.get("合约名称", "")) or (f"{underlying}{month}" if underlying else "")

    trade_unit = str(detail.get("交易单位", ""))
    min_tick = str(detail.get("最小变动价位", ""))
    margin_text = str(detail.get("最低交易保证金", ""))

    list_date = (
        normalize_date_text(base.get("上市日"))
        or normalize_date_text(base.get("开始交易日"))
        or ""
    )
    last_trade_date = (
        normalize_date_text(base.get("到期日"))
        or normalize_date_text(base.get("最后交易日"))
        or ""
    )
    near_delivery_end = normalize_date_text(base.get("最后交割日")) or ""

    margin_ratio = parse_margin_percent(margin_text)
    if not margin_ratio:
        margin_ratio = first_number(str(comm.get("保证金-买开", "")))

    return {
        "合约代码": contract,
        "合约名称": contract_name,
        "产品类型": "期货",
        "合约月份": contract_month_to_date(month),
        "衍生品代码": prod_code,
        "交易所": exchange_name,
        "币种": "CNY",
        "合约乘数": parse_multiplier(trade_unit),
        "合约上市日": list_date,
        "最后交易日": last_trade_date,
        "最小变动价格": first_number(min_tick),
        "临近交割月结束日": near_delivery_end,
        "保证金比例%": margin_ratio,
        "计量单位": parse_unit(trade_unit),
        "标的资产": underlying or contract_name,
    }


def write_to_template(rows: list[dict[str, str]], output_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"模板不存在: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]

    template_headers = [ws.cell(1, i).value or "" for i in range(1, ws.max_column + 1)]
    remove_indexes = [i for i, h in enumerate(template_headers, start=1) if h in REMOVED_HEADERS]
    for col_idx in sorted(remove_indexes, reverse=True):
        ws.delete_cols(col_idx, 1)

    actual_headers = [ws.cell(1, i).value or "" for i in range(1, ws.max_column + 1)]
    actual_headers = [h for h in actual_headers if h]
    if actual_headers != HEADERS:
        raise ValueError(f"模板表头不匹配: {actual_headers}")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, row in enumerate(rows, start=2):
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(idx, col, row.get(header, ""))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    user = read_user_input()
    print("正在通过 AKShare 拉取数据，请稍候...")

    base_info = fetch_exchange_base_info(user.query_date)
    comm_map = fetch_comm_info_map()
    detail_map = {contract: fetch_contract_detail_map(contract) for contract in user.contracts}

    missing = [c for c in user.contracts if c.lower() not in comm_map and not detail_map.get(c)]
    if missing:
        raise ValueError(f"以下合约未在 AKShare 查询到: {missing}")

    rows = [
        build_row(contract=c, base_info=base_info, comm_map=comm_map, detail_map=detail_map)
        for c in user.contracts
    ]
    output_path = build_output_path(user.output)
    write_to_template(rows, output_path)

    print(f"已生成 {len(rows)} 条合约记录: {output_path}")


if __name__ == "__main__":
    main()
