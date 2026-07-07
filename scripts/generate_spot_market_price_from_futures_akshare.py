#!/usr/bin/env python3
"""根据期货合约价格模拟现货市场价导入文件（交互式输入）。"""

from __future__ import annotations

import re
import random
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import akshare as ak
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
TEMPLATE_PATH = PROJECT_ROOT / "template" / "现货市场价导入.xlsx"

HEADERS = [
    "日期",
    "品名",
    "材质",
    "规格",
    "产地",
    "省",
    "市",
    "公允价值",
    "是否基价",
    "类型",
    "业务机构",
    "业务部门",
]


@dataclass(frozen=True)
class UserInput:
    contracts: list[str]
    names: list[str]
    material: str
    spec: str
    origin: str
    province: str
    city: str
    premium_text: str


def prompt_with_default(prompt_text: str, default: str) -> str:
    text = input(f"{prompt_text} [{default}]: ").strip()
    return text if text else default


def parse_contracts(raw: str) -> list[str]:
    items = [x.strip().upper() for x in raw.split(",") if x.strip()]
    if not items:
        raise ValueError("合约不能为空")
    pat = re.compile(r"^[A-Z]{1,2}\d{3,4}$")
    result: list[str] = []
    for item in items:
        if not pat.match(item):
            raise ValueError(f"合约格式错误: {item}，示例: RB2610")
        if item not in result:
            result.append(item)
    return result


def parse_names(raw: str, contract_count: int) -> list[str]:
    names = [x.strip() for x in raw.split(",") if x.strip()]
    if not names:
        raise ValueError("品名不能为空")
    if len(names) == 1:
        return names * contract_count
    if len(names) != contract_count:
        raise ValueError("品名数量需为 1 个或与合约数量一致")
    return names


def parse_quick_fields(raw: str) -> tuple[str, str, str, str, str, str]:
    text = raw.strip().strip("【】[]")
    if not text:
        return "", "", "", "", "", ""

    # 支持 tab 或英文逗号分隔:
    # 1) 品名,材质,规格,产地,省,市
    # 2) 材质,规格,产地,省,市
    if "\t" in text:
        parts = [x.strip() for x in text.split("\t") if x.strip()]
    else:
        parts = [x.strip() for x in text.split(",") if x.strip()]

    if len(parts) == 6:
        return parts[0], parts[1], parts[2], parts[3], parts[4], parts[5]
    if len(parts) == 5:
        return "", parts[0], parts[1], parts[2], parts[3], parts[4]
    raise ValueError("一次性输入格式错误，应为 5 段或 6 段")


def read_user_input() -> UserInput:
    print("请输入参数（直接回车使用默认值）")
    contracts_text = prompt_with_default("合约代码(英文逗号分隔)", "RB2610")
    contracts = parse_contracts(contracts_text)

    names_text = prompt_with_default("品名(1个或与合约数量一致，英文逗号分隔)", "螺纹钢现货")
    names = parse_names(names_text, len(contracts))

    quick_text = input(
        "一次性输入[品名\\t材质\\t规格\\t产地\\t省\\t市] (可回车跳过): "
    ).strip()
    if quick_text:
        quick_name, material, spec, origin, province, city = parse_quick_fields(quick_text)
        if quick_name:
            names = [quick_name] * len(contracts)
    else:
        material = input("材质(可回车为空): ").strip()
        spec = input("规格(可回车为空): ").strip()
        origin = input("产地(可回车为空): ").strip()
        province = input("省(可回车为空): ").strip()
        city = input("市(可回车为空): ").strip()
    premium_text = input("升贴水(固定值或百分比如1%，回车默认按合约价随机[-1%,1%]): ").strip()

    return UserInput(
        contracts=contracts,
        names=names,
        material=material,
        spec=spec,
        origin=origin,
        province=province,
        city=city,
        premium_text=premium_text,
    )


def build_output_path() -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return PROJECT_ROOT / "dist" / f"现货市场价导入_{ts}.xlsx"


def to_float(value: object) -> float | None:
    if pd.isna(value):
        return None
    try:
        return float(value)
    except Exception:
        return None


def calc_premium(settle_price: float, premium_text: str) -> float:
    if not premium_text:
        return settle_price * random.uniform(-0.01, 0.01)

    text = premium_text.replace(" ", "")
    if text.endswith("%"):
        ratio = abs(float(text[:-1]) / 100.0)
        return settle_price * random.uniform(-ratio, ratio)

    amount = abs(float(text))
    return random.uniform(-amount, amount)


def fetch_spot_rows(contract: str, name: str, user: UserInput, start_date: date, end_date: date) -> list[dict[str, str]]:
    df = ak.futures_zh_daily_sina(symbol=contract)
    if df.empty:
        return []

    work = df.copy()
    work["date"] = pd.to_datetime(work["date"])
    work = work.sort_values("date")

    # 自然日补全: 周末/节假日无结算价时，向前取最近一个可用结算价
    calendar = pd.DataFrame(
        {
            "date": pd.date_range(
                start=pd.Timestamp(start_date),
                end=pd.Timestamp(end_date),
                freq="D",
            )
        }
    )
    joined = pd.merge_asof(
        calendar,
        work[["date", "settle"]],
        on="date",
        direction="backward",
    )
    joined = joined[joined["settle"].notna()].copy()
    joined["date"] = joined["date"].dt.date

    rows: list[dict[str, str]] = []
    last_fair_value: float | None = None
    for idx, row in joined.iterrows():
        settle = to_float(row.get("settle"))
        if settle is None:
            continue

        premium = calc_premium(settle, user.premium_text)
        fair_value = settle + premium
        # 尽量保证每天值不重复: 若与前一天相同，则做最小价差微调
        if last_fair_value is not None and round(fair_value, 2) == round(last_fair_value, 2):
            fair_value += 0.01 * ((idx % 3) + 1)
        last_fair_value = fair_value

        rows.append(
            {
                "日期": row.get("date").strftime("%Y-%m-%d"),
                "品名": name,
                "材质": user.material,
                "规格": user.spec,
                "产地": user.origin,
                "省": user.province,
                "市": user.city,
                "公允价值": f"{fair_value:.2f}",
                "是否基价": "√",
                "类型": "",
                "业务机构": "",
                "业务部门": "",
            }
        )

    return rows


def write_to_template(rows: list[dict[str, str]], output_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"模板不存在: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]

    actual_headers = [ws.cell(1, i).value or "" for i in range(1, len(HEADERS) + 1)]
    if actual_headers != HEADERS:
        raise ValueError(f"模板表头不匹配: {actual_headers}")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            ws.cell(r_idx, c_idx, row.get(header, ""))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    user = read_user_input()

    end_date = date.today()
    start_date = end_date - timedelta(days=365)
    print(f"时间范围固定为: {start_date} ~ {end_date}")
    print("正在通过 AKShare 拉取合约价格并模拟现货价格，请稍候...")

    all_rows: list[dict[str, str]] = []
    skipped: list[str] = []

    for contract, name in zip(user.contracts, user.names):
        rows = fetch_spot_rows(contract, name, user, start_date, end_date)
        if not rows:
            skipped.append(contract)
            continue
        all_rows.extend(rows)

    if not all_rows:
        raise ValueError("没有可生成的数据，请检查合约是否有效")

    all_rows.sort(key=lambda x: (x["日期"], x["品名"]))
    output_path = build_output_path()
    write_to_template(all_rows, output_path)

    print(f"已生成 {len(all_rows)} 条现货市场价记录: {output_path}")
    if skipped:
        print(f"以下合约无数据已跳过: {', '.join(skipped)}")


if __name__ == "__main__":
    main()
